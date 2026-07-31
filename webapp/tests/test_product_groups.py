from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from webapp.exporter import build_export
from webapp.product_groups import (
    aggregate_product_groups,
    canonical_msku,
)
from webapp.repository import Repository
from webapp.service import _inbound_index, build_dashboard, build_product_detail


def grouped_snapshot():
    return {
        "collected_at": datetime.now().isoformat(timespec="seconds"),
        "source_date": "2026-07-29",
        "source": "测试",
        "stores": [{"store_id": "STORE-1", "store_name": "测试店铺"}],
        "products": [
            {
                "product_name": "测试商品旧编码",
                "msku": "PAIR-001",
                "sku": "PAIR-001",
                "store_id": "STORE-1",
                "store_name": "测试店铺",
                "avg_7": 2,
                "avg_14": 1,
                "avg_30": 0,
                "fbt_total": 10,
                "fbt_sellable": 8,
                "fbt_in_transit": 5,
            },
            {
                "product_name": "测试商品",
                "msku": "PAIR-001-US",
                "sku": "PAIR-001",
                "store_id": "STORE-1",
                "store_name": "测试店铺",
                "avg_7": 3,
                "avg_14": 2,
                "avg_30": 1,
                "fbt_total": 20,
                "fbt_sellable": 15,
                "fbt_in_transit": 7,
            },
        ],
        "shipments": [],
        "shipping_orders": [],
    }


def test_canonical_msku_only_removes_the_terminal_us_suffix():
    assert canonical_msku("PAIR-001-US") == "PAIR-001"
    assert canonical_msku("PAIR-US-001") == "PAIR-US-001"
    assert canonical_msku("PAIR-001-FBT") == "PAIR-001-FBT"


def test_group_aggregates_sales_and_inventory_before_calculation():
    item = aggregate_product_groups(grouped_snapshot()["products"])[0]

    assert item["canonical_msku"] == "PAIR-001"
    assert item["execution_msku"] == "PAIR-001-US"
    assert item["member_mskus"] == ["PAIR-001", "PAIR-001-US"]
    assert item["avg_7"] == 5
    assert item["avg_14"] == 3
    assert item["avg_30"] == 1
    assert item["fbt_total"] == 30
    assert item["fbt_sellable"] == 23
    assert item["fbt_in_transit"] == 12
    assert item["fbt_all"] == 42


def test_same_canonical_msku_in_different_stores_stays_separate():
    products = grouped_snapshot()["products"]
    products.append(
        {
            **products[0],
            "store_id": "STORE-2",
            "store_name": "第二店铺",
        }
    )

    groups = aggregate_product_groups(products)

    assert len(groups) == 2
    assert {item["store_id"] for item in groups} == {"STORE-1", "STORE-2"}


def test_dashboard_calculates_one_group_and_detail_accepts_either_alias(tmp_path):
    repository = Repository(tmp_path / "test.db")
    repository.save_snapshot(grouped_snapshot())

    dashboard = build_dashboard(repository, "2026-07-29")
    item = dashboard["products"][0]

    assert dashboard["snapshot"]["raw_product_count"] == 2
    assert dashboard["snapshot"]["product_count"] == 1
    assert item["dynamic_daily"] == 3.6
    assert item["inventory_position"] == 42
    assert item["is_grouped"] is True
    assert build_product_detail(
        repository, "PAIR-001", "STORE-1", "2026-07-29"
    )["product"]["product_group_id"] == item["product_group_id"]
    assert build_product_detail(
        repository, "PAIR-001-US", "STORE-1", "2026-07-29"
    )["product"]["product_group_id"] == item["product_group_id"]


def test_shipping_export_contains_one_row_for_the_merged_group(tmp_path):
    repository = Repository(tmp_path / "test.db")
    repository.save_snapshot(grouped_snapshot())
    dashboard = build_dashboard(repository, "2026-07-29")

    workbook = load_workbook(BytesIO(build_export(dashboard)))
    sheet = workbook.active

    assert sheet.max_row == 2
    assert sheet["B2"].value == "PAIR-001-US"
    assert sheet["C2"].value == 5
    assert sheet["G2"].value == 30


def test_manual_execution_msku_and_group_status_are_shared(tmp_path):
    repository = Repository(tmp_path / "test.db")
    repository.save_snapshot(grouped_snapshot())
    repository.save_product_group_execution(
        "STORE-1", "PAIR-001", "PAIR-001"
    )
    repository.save_product_planning_status(
        "STORE-1", "PAIR-001-US", "clearance"
    )

    item = build_dashboard(repository, "2026-07-29")["products"][0]

    assert item["execution_msku"] == "PAIR-001"
    assert item["planning_status"] == "clearance"
    assert item["is_planning_excluded"] is True


def test_inbound_items_for_aliases_enter_one_group_ledger():
    shipments = [
        {
            "is_received": False,
            "is_archived": False,
            "is_api_synced": True,
            "order_status_name": "",
            "ship_status": "",
            "shipping_list_code": "LIST-1",
            "delivery_time": "2026-07-20",
            "expected_delivery_date": "2026-08-20",
            "tracking_number": "TRACK-1",
            "carrier": "UPS",
            "manual_expected_delivery_date": None,
            "status": "in_transit",
            "items": [
                {
                    "cargo_code": "IBR-1",
                    "store_id": "STORE-1",
                    "msku": "PAIR-001",
                    "remaining_qty": 10,
                },
                {
                    "cargo_code": "IBR-1",
                    "store_id": "STORE-1",
                    "msku": "PAIR-001-US",
                    "remaining_qty": 20,
                },
            ],
        }
    ]

    index = _inbound_index(shipments)

    assert list(index) == [("STORE-1", "PAIR-001")]
    assert sum(item["remaining_qty"] for item in index[("STORE-1", "PAIR-001")]) == 30
