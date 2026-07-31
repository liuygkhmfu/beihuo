from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from webapp.exporter import build_purchase_export
from webapp.purchase import (
    build_purchase_plan,
    remaining_equivalent_days,
    round_purchase_daily,
)
from webapp.repository import Repository


def purchase_snapshot():
    return {
        "collected_at": datetime.now().isoformat(timespec="seconds"),
        "source_date": "2026-07-29",
        "source": "测试",
        "stores": [{"store_id": "STORE-1", "store_name": "测试店铺"}],
        "products": [
            {
                "product_name": "测试商品",
                "msku": "TEST-001-US",
                "sku": "TEST-001",
                "store_id": "STORE-1",
                "store_name": "测试店铺",
                "avg_7": 10,
                "avg_14": 8,
                "avg_30": 6,
                "fbt_total": 9999,
                "fbt_sellable": 9999,
                "fbt_in_transit": 9999,
            }
        ],
        "shipments": [],
        "shipping_orders": [],
    }


def test_remaining_days_are_cumulative_not_single_month():
    assert remaining_equivalent_days(0) == 318.75
    assert remaining_equivalent_days(7) == 288.75
    assert remaining_equivalent_days(8) == 258.75
    assert remaining_equivalent_days(9) == 213.75
    assert remaining_equivalent_days(10) == 168.75
    assert remaining_equivalent_days(11) == 101.25
    assert remaining_equivalent_days(12) == 0


def test_purchase_daily_uses_half_up_rounding_to_two_decimals():
    assert round_purchase_daily(1.234) == 1.23
    assert round_purchase_daily(1.235) == 1.24
    assert round_purchase_daily(0.004) == 0
    assert round_purchase_daily(0.005) == 0.01


def test_purchase_plan_hides_items_rounded_to_zero(tmp_path):
    repository = Repository(tmp_path / "test.db")
    snapshot = purchase_snapshot()
    snapshot["products"].append(
        {
            "product_name": "LOW-SALES",
            "msku": "LOW-001-US",
            "sku": "LOW-001",
            "store_id": "STORE-1",
            "store_name": "STORE-1",
            "avg_7": 0.004,
            "avg_14": 0.004,
            "avg_30": 0.004,
            "fbt_total": 0,
            "fbt_sellable": 0,
            "fbt_in_transit": 0,
        }
    )
    repository.save_snapshot(snapshot)
    plan = build_purchase_plan(repository, "2026-07-29")
    assert [item["sku"] for item in plan["items"]] == ["TEST-001"]


def test_purchase_plan_excludes_clearance_products(tmp_path):
    repository = Repository(tmp_path / "test.db")
    repository.save_snapshot(purchase_snapshot())
    repository.save_product_planning_status(
        "STORE-1", "TEST-001-US", "clearance"
    )

    plan = build_purchase_plan(repository, "2026-07-29")

    assert plan["items"] == []
    assert plan["summary"]["final_qty_total"] == 0


def test_purchase_plan_uses_daily_times_remaining_days_without_inventory_deduction(
    tmp_path,
):
    repository = Repository(tmp_path / "test.db")
    repository.save_snapshot(purchase_snapshot())
    plan = build_purchase_plan(repository, "2026-07-29")
    item = plan["items"][0]
    assert plan["completed_month"] == 7
    assert plan["remaining_equivalent_days"] == 288.75
    assert item["dynamic_daily"] == 8.6
    assert item["system_qty"] == 2484


def test_purchase_overrides_adjust_daily_days_and_final_quantity(tmp_path):
    repository = Repository(tmp_path / "test.db")
    repository.save_snapshot(purchase_snapshot())
    repository.save_purchase_plan_config(2026, 7)
    repository.save_purchase_plan_overrides(
        2026,
        [
            {
                "sku_key": "TEST-001",
                "adopted_daily": 10,
                "extra_days": 15,
                "final_qty": 3000,
                "note": "达人重点款",
            }
        ],
    )
    item = build_purchase_plan(repository, "2026-07-29")["items"][0]
    assert item["system_qty"] == 3038
    assert item["final_qty"] == 3000
    assert item["has_manual_adjustment"] is True
    repository.save_purchase_plan_overrides(
        2026,
        [
            {
                "sku_key": "TEST-001",
                "adopted_daily": None,
                "extra_days": 0,
                "final_qty": None,
                "note": "",
            }
        ],
    )
    assert repository.get_purchase_plan_overrides(2026) == {}


def test_purchase_export_contains_only_three_supplier_columns(tmp_path):
    repository = Repository(tmp_path / "test.db")
    repository.save_snapshot(purchase_snapshot())
    body = build_purchase_export(build_purchase_plan(repository, "2026-07-29"))
    workbook = load_workbook(BytesIO(body), data_only=False)
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == ["SKU", "品名", "备货量"]
    assert sheet.max_column == 3
    assert sheet["A2"].value == "TEST-001"
    assert sheet["B2"].value == "测试商品"
    assert sheet["C2"].value == 2484
