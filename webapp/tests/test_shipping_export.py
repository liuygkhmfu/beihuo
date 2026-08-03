from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from webapp.domain import (
    DEFAULT_SCHEDULE,
    DEFAULT_SETTINGS,
    calculate_recommendation,
)
from webapp.exporter import BASE_SHIPPING_HEADERS, build_export


def _dashboard(**setting_overrides):
    settings = {**DEFAULT_SETTINGS, **setting_overrides}
    product = {
        "product_name": "测试商品",
        "msku": "TEST-001",
        "store_id": "STORE-1",
        "store_name": "测试店铺",
        "avg_7": 10,
        "avg_14": 8,
        "avg_30": 6,
        "fbt_total": 100,
        "fbt_sellable": 80,
        "fbt_in_transit": 50,
    }
    recommendation = calculate_recommendation(
        product,
        settings,
        DEFAULT_SCHEDULE,
        date(2026, 7, 28),
    )
    return {"settings": settings, "products": [recommendation]}


def _headers_and_row(dashboard):
    workbook = load_workbook(BytesIO(build_export(dashboard)))
    sheet = workbook["补货建议"]
    headers = [cell.value for cell in sheet[1]]
    row = [cell.value for cell in sheet[2]]
    return workbook.sheetnames, headers, row


def test_shipping_export_is_a_single_simple_data_sheet():
    sheetnames, headers, row = _headers_and_row(_dashboard())

    assert sheetnames == ["补货建议"]
    assert headers == [
        *BASE_SHIPPING_HEADERS,
        "快船",
        "普船卡派",
        "COSCO慢船",
        "快船建议补货",
        "普船卡派建议补货",
        "COSCO慢船建议补货",
    ]
    assert row[0:3] == ["测试商品", "TEST-001", 10]
    assert row[headers.index("快船")] == 41
    assert row[headers.index("COSCO慢船")] == 130.75


def test_shipping_export_channel_columns_follow_channel_switches():
    _, headers, _ = _headers_and_row(
        _dashboard(
            air_channel_enabled=True,
            truck_channel_enabled=False,
        )
    )

    assert headers == [
        *BASE_SHIPPING_HEADERS,
        "空派 IE",
        "快船",
        "COSCO慢船",
        "空派 IE建议补货",
        "快船建议补货",
        "COSCO慢船建议补货",
    ]
    assert "普船卡派" not in headers


def test_shipping_export_lists_express_and_air_as_separate_channels():
    _, headers, _ = _headers_and_row(
        _dashboard(
            express_channel_enabled=True,
            air_channel_enabled=True,
        )
    )

    assert "快递 IP" in headers
    assert "空派 IE" in headers
    assert "快递 IP建议补货" in headers
    assert "空派 IE建议补货" in headers
    assert headers.index("快递 IP") < headers.index("空派 IE")


def test_shipping_export_excludes_clearance_and_delisted_products():
    dashboard = _dashboard()
    dashboard["products"][0]["is_planning_excluded"] = True

    workbook = load_workbook(BytesIO(build_export(dashboard)))
    sheet = workbook["补货建议"]

    assert sheet.max_row == 1


def test_shipping_export_uses_the_formally_effective_reviewed_quantities():
    dashboard = _dashboard()
    item = dashboard["products"][0]
    channels = [
        plan for plan in item["channel_plans"] if plan.get("enabled", True)
    ]
    expected = []
    for index, channel in enumerate(channels, start=1):
        quantity = index * 111
        item[f"effective_{channel['key']}_qty"] = quantity
        expected.append(quantity)

    _, _, row = _headers_and_row(dashboard)
    quantity_start = len(BASE_SHIPPING_HEADERS) + len(channels)

    assert row[quantity_start:] == expected
