from datetime import datetime
from copy import deepcopy
from io import BytesIO

from openpyxl import Workbook

from webapp.arrival_tracking import import_arrival_workbook, parse_arrival_workbook
from webapp.exporter import build_arrival_tracking_export
from webapp.repository import Repository
from webapp.service import build_shipments


def arrival_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = datetime(2026, 7, 23)
    sheet.merge_cells("A1:A2")
    sheet["B1"] = "第34批已发货 IBR5767694851696726022"
    sheet.merge_cells("B1:B2")
    sheet["C1"] = "DM768-3"
    sheet["D1"] = "慢回弹小海龟菠萝包"
    sheet["E1"] = 100
    sheet["C2"] = "OTHER-1"
    sheet["D2"] = "其他商品"
    sheet["E2"] = 20
    sheet["I1"] = "预计7.29签收；"
    sheet.merge_cells("I1:I2")
    sheet["J1"] = "7.23开船，7.30到港"
    sheet.merge_cells("J1:J2")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def shipment_snapshot() -> dict:
    return {
        "source": "test",
        "source_date": "2026-07-29",
        "collected_at": "2026-07-29T10:00:00",
        "stores": [{"store_id": "STORE-1", "store_name": "TK US"}],
        "products": [
            {
                "store_id": "STORE-1",
                "store_name": "TK US",
                "msku": "DM768-3-US",
                "sku": "DM768-3",
                "product_name": "慢回弹小海龟菠萝包",
            }
        ],
        "shipments": [
            {
                "cargo_id": "CARGO-1",
                "cargo_code": "IBR5767694851696726022",
                "store_id": "STORE-1",
                "store_name": "TK US",
                "order_status": "2",
                "order_status_name": "已发货",
                "ship_status": "运输中",
                "shipping_warehouse": "FBT",
                "create_time": "2026-07-23",
                "delivery_time": "2026-07-23",
                "expected_delivery_time": "2026-08-05",
                "actual_delivery_time": None,
                "shipping_list_codes": [],
                "items": [
                    {
                        "msku": "DM768-3-US",
                        "sku": "DM768-3",
                        "product_name": "慢回弹小海龟菠萝包",
                        "image_url": "",
                        "declaration_qty": 100,
                        "shipment_qty": 100,
                        "signed_qty": 100,
                        "normal_qty": 40,
                        "defective_qty": 0,
                    }
                ],
            }
        ],
        "shipping_orders": [],
    }


def test_import_matches_platform_sku_and_is_idempotent(tmp_path):
    repository = Repository(tmp_path / "test.db")
    repository.save_snapshot(shipment_snapshot())
    content = arrival_workbook_bytes()

    imported = import_arrival_workbook(
        repository, content, "到货跟踪.xlsx", "STORE-1"
    )
    duplicate = import_arrival_workbook(
        repository, content, "到货跟踪.xlsx", "STORE-1"
    )

    assert imported["matched_count"] == 1
    assert imported["unmatched_count"] == 1
    assert duplicate["duplicate"] is True
    batches = repository.get_arrival_batches()
    matched = next(
        item for item in batches[0]["items"] if item["raw_sku"] == "DM768-3"
    )
    assert matched["matched_msku"] == "DM768-3-US"
    assert matched["match_method"] == "SKU完全一致"


def test_signed_quantity_does_not_equal_fbt_received_quantity(tmp_path):
    repository = Repository(tmp_path / "test.db")
    repository.save_snapshot(shipment_snapshot())
    import_arrival_workbook(
        repository, arrival_workbook_bytes(), "到货跟踪.xlsx", "STORE-1"
    )

    shipment = build_shipments(repository, "2026-07-29")["shipments"][0]

    assert shipment["signed_qty"] == 100
    assert shipment["received_qty"] == 40
    assert shipment["awaiting_receive_qty"] == 60
    assert shipment["remaining_qty"] == 60
    assert shipment["expected_signed_date"] == "2026-07-29"
    assert shipment["expected_receive_date"] == "2026-07-30"


def test_impossible_api_eta_before_departure_is_not_used(tmp_path):
    repository = Repository(tmp_path / "test.db")
    snapshot = shipment_snapshot()
    snapshot["shipments"][0]["expected_delivery_time"] = "2026-07-20"
    repository.save_snapshot(snapshot)

    shipment = build_shipments(repository, "2026-07-29")["shipments"][0]

    assert shipment["departure_date"] == "2026-07-23"
    assert shipment["expected_receive_date"] is None


def test_api_shipments_before_tracking_baseline_are_archived(tmp_path):
    repository = Repository(tmp_path / "test.db")
    snapshot = shipment_snapshot()
    old_shipment = deepcopy(snapshot["shipments"][0])
    old_shipment.update(
        {
            "cargo_id": "OLD-CARGO",
            "cargo_code": "IBR-OLD",
            "create_time": "2026-07-01",
            "delivery_time": "2026-07-01",
            "expected_delivery_time": "2026-07-20",
        }
    )
    snapshot["shipments"].append(old_shipment)
    repository.save_snapshot(snapshot)
    import_arrival_workbook(
        repository, arrival_workbook_bytes(), "到货跟踪.xlsx", "STORE-1"
    )

    shipment_data = build_shipments(repository, "2026-07-29")
    old_view = next(
        item for item in shipment_data["shipments"] if item["cargo_code"] == "IBR-OLD"
    )

    assert old_view["status"] == "archived"
    assert old_view["is_overdue"] is False
    assert shipment_data["summary"]["archived_count"] == 1
    assert shipment_data["summary"]["active_count"] == 1


def test_exported_workbook_can_be_imported_again(tmp_path):
    repository = Repository(tmp_path / "test.db")
    repository.save_snapshot(shipment_snapshot())
    import_arrival_workbook(
        repository, arrival_workbook_bytes(), "到货跟踪.xlsx", "STORE-1"
    )
    shipment_data = build_shipments(repository, "2026-07-29")

    exported = build_arrival_tracking_export(shipment_data)
    batches = parse_arrival_workbook(exported, "STORE-1")

    assert len(batches) == 1
    assert batches[0]["cargo_code"] == "IBR5767694851696726022"
    assert any(
        item["matched_msku"] == "DM768-3-US" for item in batches[0]["items"]
    )
