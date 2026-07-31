from __future__ import annotations

import hashlib
import re
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .product_groups import canonical_msku, select_execution_product


IBR_PATTERN = re.compile(r"IBR\d+", re.IGNORECASE)
SKU_SUFFIX_PATTERN = re.compile(r"(?:[-_\s]+(?:US|FBA|FBT))+$", re.IGNORECASE)
DATE_BEFORE_KEYWORD = (
    r"(?P<month>\d{1,2})\s*[./月]\s*(?P<day>\d{1,3})\s*(?:日)?\s*"
)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _date_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and value > 1000:
        try:
            return from_excel(value).date().isoformat()
        except (TypeError, ValueError, OverflowError):
            return None
    text = _text(value)
    if not text:
        return None
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return None


def _extract_ibr(value: Any) -> str:
    match = IBR_PATTERN.search(_text(value))
    return match.group(0).upper() if match else ""


def _year_for_month(year_hint: int, month_hint: int, month: int) -> int:
    if month_hint >= 10 and month <= 3:
        return year_hint + 1
    if month_hint <= 3 and month >= 10:
        return year_hint - 1
    return year_hint


def _extract_event_date(
    value: Any,
    keyword: str,
    year_hint: int,
    month_hint: int,
    prefix: str = "",
) -> str | None:
    text = _text(value)
    if not text:
        return None
    pattern = (
        (re.escape(prefix) if prefix else "")
        + DATE_BEFORE_KEYWORD
        + re.escape(keyword)
    )
    matches = list(re.finditer(pattern, text))
    if not matches:
        return None
    match = matches[-1]
    month = int(match.group("month"))
    day = int(match.group("day"))
    year = _year_for_month(year_hint, month_hint, month)
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return None


def _normalize_sku(value: Any) -> str:
    text = re.sub(r"\s+", " ", _text(value).upper()).strip()
    return SKU_SUFFIX_PATTERN.sub("", text)


def _stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join(_text(part) for part in parts)
    return f"{prefix}_{hashlib.sha1(raw.encode('utf-8')).hexdigest()[:20]}"


def _merged_value_map(sheet: Any) -> dict[tuple[int, int], Any]:
    result: dict[tuple[int, int], Any] = {}
    for merged in sheet.merged_cells.ranges:
        value = sheet.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                result[(row, column)] = value
    return result


def _legacy_workbook_rows(workbook: Any) -> list[dict[str, Any]]:
    sheet = workbook[workbook.sheetnames[0]]
    merged_values = _merged_value_map(sheet)
    rows: list[dict[str, Any]] = []
    carry: dict[int, Any] = {}
    for row_number in range(1, sheet.max_row + 1):
        values = []
        for column in range(1, 11):
            value = merged_values.get(
                (row_number, column), sheet.cell(row_number, column).value
            )
            if column in {1, 2, 6, 7, 8, 9, 10}:
                if value not in (None, ""):
                    carry[column] = value
                elif column in {1, 2}:
                    value = carry.get(column)
            values.append(value)
        raw_sku = _text(values[2])
        quantity = _number(values[4])
        if not raw_sku or quantity <= 0:
            continue
        rows.append(
            {
                "source_row": row_number,
                "shipment_date": _date_text(values[0]),
                "batch_label": _text(values[1]),
                "raw_sku": raw_sku,
                "product_name": _text(values[3]),
                "shipment_qty": quantity,
                "metric_f": values[5],
                "metric_g": values[6],
                "metric_h": values[7],
                "status_note": _text(values[8]),
                "route_note": _text(values[9]),
            }
        )
    return rows


def _system_workbook_rows(workbook: Any) -> list[dict[str, Any]]:
    sheet = workbook["_系统数据"]
    headers = [_text(cell.value) for cell in sheet[1]]
    result = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = dict(zip(headers, row))
        if not _text(values.get("item_id")):
            continue
        result.append(values)
    return result


def _rows_to_batches(
    rows: list[dict[str, Any]], default_store_id: str
) -> list[dict[str, Any]]:
    if rows and "batch_id" in rows[0]:
        batches: dict[str, dict[str, Any]] = {}
        for row in rows:
            batch_id = _text(row.get("batch_id"))
            if not batch_id:
                continue
            batch = batches.setdefault(
                batch_id,
                {
                    "id": batch_id,
                    "cargo_code": _text(row.get("cargo_code")),
                    "store_id": _text(row.get("store_id")) or default_store_id,
                    "batch_label": _text(row.get("batch_label")),
                    "shipment_date": _date_text(row.get("shipment_date")),
                    "departure_date": _date_text(row.get("departure_date")),
                    "port_arrival_date": _date_text(row.get("port_arrival_date")),
                    "expected_signed_date": _date_text(
                        row.get("expected_signed_date")
                    ),
                    "actual_signed_date": _date_text(row.get("actual_signed_date")),
                    "expected_receive_date": _date_text(
                        row.get("expected_receive_date")
                    ),
                    "actual_receive_date": _date_text(
                        row.get("actual_receive_date")
                    ),
                    "is_fully_received": bool(row.get("is_fully_received")),
                    "carrier": _text(row.get("carrier")),
                    "tracking_number": _text(row.get("tracking_number")),
                    "status_note": _text(row.get("status_note")),
                    "route_note": _text(row.get("route_note")),
                    "metric_f": row.get("metric_f"),
                    "metric_g": row.get("metric_g"),
                    "metric_h": row.get("metric_h"),
                    "source_row_start": int(row.get("source_row") or 0),
                    "source_row_end": int(row.get("source_row") or 0),
                    "items": [],
                },
            )
            batch["items"].append(
                {
                    "id": _text(row.get("item_id")),
                    "source_row": int(row.get("source_row") or 0),
                    "raw_sku": _text(row.get("raw_sku")),
                    "matched_store_id": _text(row.get("matched_store_id")),
                    "matched_msku": _text(row.get("matched_msku")),
                    "matched_sku": _text(row.get("matched_sku")),
                    "product_name": _text(row.get("product_name")),
                    "shipment_qty": _number(row.get("shipment_qty")),
                    "match_status": _text(row.get("match_status")) or "unmatched",
                    "match_method": _text(row.get("match_method")),
                    "conflict_note": _text(row.get("conflict_note")),
                }
            )
        return list(batches.values())

    grouped: dict[str, dict[str, Any]] = {}
    occurrences: dict[tuple[str, str], int] = defaultdict(int)
    for row in rows:
        cargo_code = _extract_ibr(row["batch_label"])
        batch_key = cargo_code or _stable_id(
            "manual",
            row.get("shipment_date"),
            row.get("batch_label"),
        )
        if cargo_code:
            batch_id = f"ibr_{cargo_code}"
        else:
            batch_id = batch_key
        shipment_date = _date_text(row.get("shipment_date"))
        date_hint = date.fromisoformat(shipment_date) if shipment_date else date.today()
        status_note = _text(row.get("status_note"))
        route_note = _text(row.get("route_note"))
        batch = grouped.setdefault(
            batch_id,
            {
                "id": batch_id,
                "cargo_code": cargo_code,
                "store_id": default_store_id,
                "batch_label": _text(row.get("batch_label")),
                "shipment_date": shipment_date,
                "departure_date": _extract_event_date(
                    route_note, "开船", date_hint.year, date_hint.month
                ),
                "port_arrival_date": _extract_event_date(
                    route_note, "到港", date_hint.year, date_hint.month
                ),
                "expected_signed_date": _extract_event_date(
                    status_note, "签收", date_hint.year, date_hint.month, "预计"
                ),
                "actual_signed_date": _extract_event_date(
                    status_note, "已签收", date_hint.year, date_hint.month
                ),
                "expected_receive_date": None,
                "actual_receive_date": _extract_event_date(
                    status_note, "已全部接收", date_hint.year, date_hint.month
                ),
                "is_fully_received": "已全部接收" in status_note,
                "carrier": "",
                "tracking_number": "",
                "status_note": status_note,
                "route_note": route_note,
                "metric_f": row.get("metric_f"),
                "metric_g": row.get("metric_g"),
                "metric_h": row.get("metric_h"),
                "source_row_start": int(row["source_row"]),
                "source_row_end": int(row["source_row"]),
                "items": [],
            },
        )
        batch["source_row_start"] = min(
            batch["source_row_start"], int(row["source_row"])
        )
        batch["source_row_end"] = max(batch["source_row_end"], int(row["source_row"]))
        normalized = _normalize_sku(row["raw_sku"])
        occurrence_key = (batch_id, normalized)
        occurrences[occurrence_key] += 1
        batch["items"].append(
            {
                "id": _stable_id(
                    "item", batch_id, normalized, occurrences[occurrence_key]
                ),
                "source_row": int(row["source_row"]),
                "raw_sku": row["raw_sku"],
                "matched_store_id": "",
                "matched_msku": "",
                "matched_sku": "",
                "product_name": row["product_name"],
                "shipment_qty": row["shipment_qty"],
                "match_status": "unmatched",
                "match_method": "",
                "conflict_note": "",
            }
        )
    return list(grouped.values())


def parse_arrival_workbook(
    content: bytes, default_store_id: str = ""
) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(f"无法读取到货跟踪表：{exc}") from exc
    if "_系统数据" in workbook.sheetnames:
        rows = _system_workbook_rows(workbook)
    else:
        rows = _legacy_workbook_rows(workbook)
    if not rows:
        raise ValueError("到货跟踪表中没有识别到SKU和发货数量")
    return _rows_to_batches(rows, default_store_id)


def _candidate_key(candidate: dict[str, Any]) -> tuple[str, str]:
    return (
        _text(candidate.get("store_id")),
        canonical_msku(candidate.get("msku")).upper(),
    )


def _unique_candidates(
    candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    result: dict[tuple[str, str], dict[str, Any]] = {}
    for candidate in candidates:
        key = _candidate_key(candidate)
        existing = result.get(key)
        result[key] = (
            select_execution_product([existing, candidate])
            if existing
            else candidate
        )
    return list(result.values())


def _match_item(
    item: dict[str, Any],
    candidates: list[dict[str, Any]],
    aliases: dict[tuple[str, str], dict[str, Any]],
    store_id: str,
) -> dict[str, Any]:
    raw_sku = _text(item.get("raw_sku"))
    raw_upper = raw_sku.upper()
    alias = aliases.get((store_id, raw_upper)) or aliases.get(("", raw_upper))
    if alias:
        item.update(
            {
                "matched_store_id": alias["store_id"] or store_id,
                "matched_msku": alias["canonical_msku"],
                "matched_sku": alias["canonical_sku"],
                "match_status": "matched",
                "match_method": "人工别名",
                "conflict_note": "",
            }
        )
        return item

    scoped = [
        candidate
        for candidate in candidates
        if not store_id or _text(candidate.get("store_id")) == store_id
    ]
    checks = [
        (
            "MSKU完全一致",
            [
                candidate
                for candidate in scoped
                if _text(candidate.get("msku")).upper() == raw_upper
            ],
        ),
        (
            "SKU完全一致",
            [
                candidate
                for candidate in scoped
                if _text(candidate.get("sku")).upper() == raw_upper
            ],
        ),
        (
            "去除店铺后缀后唯一",
            [
                candidate
                for candidate in scoped
                if _normalize_sku(candidate.get("msku")) == _normalize_sku(raw_sku)
                or _normalize_sku(candidate.get("sku")) == _normalize_sku(raw_sku)
            ],
        ),
    ]
    for method, matches in checks:
        unique = _unique_candidates(matches)
        if len(unique) == 1:
            candidate = unique[0]
            item.update(
                {
                    "matched_store_id": _text(candidate.get("store_id")),
                    "matched_msku": _text(candidate.get("msku")),
                    "matched_sku": _text(candidate.get("sku")),
                    "product_name": item.get("product_name")
                    or _text(candidate.get("product_name")),
                    "match_status": "matched",
                    "match_method": method,
                    "conflict_note": "",
                }
            )
            return item
        if len(unique) > 1:
            item.update(
                {
                    "match_status": "conflict",
                    "match_method": method,
                    "conflict_note": f"找到{len(unique)}个候选商品，需要人工确认",
                }
            )
            return item

    same_name = _unique_candidates(
        [
            candidate
            for candidate in scoped
            if _text(candidate.get("product_name"))
            and _text(candidate.get("product_name")) == _text(item.get("product_name"))
        ]
    )
    suggestion = ""
    if len(same_name) == 1:
        suggestion = f"产品名相同，建议核对 {same_name[0].get('msku')}"
    item.update(
        {
            "match_status": "unmatched",
            "match_method": "",
            "conflict_note": suggestion or "未找到唯一SKU，请人工关联",
        }
    )
    return item


def reconcile_batches(
    repository: Any, batches: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    aliases = repository.get_product_aliases()
    catalog = repository.get_product_match_catalog()
    cargo_context = repository.get_shipment_match_context()
    for batch in batches:
        cargo_code = batch.get("cargo_code", "")
        context = cargo_context.get(cargo_code, {})
        if context.get("store_id"):
            batch["store_id"] = context["store_id"]
        cargo_candidates = context.get("items") or catalog
        for item in batch["items"]:
            _match_item(
                item,
                cargo_candidates,
                aliases,
                batch.get("store_id", ""),
            )
    return batches


def import_arrival_workbook(
    repository: Any,
    content: bytes,
    filename: str,
    default_store_id: str = "",
) -> dict[str, Any]:
    file_hash = hashlib.sha256(content).hexdigest()
    existing = repository.get_arrival_import_by_hash(file_hash)
    if existing:
        return {
            **existing,
            "duplicate": True,
            "message": "该文件已经导入，本次没有重复新增",
        }
    batches = parse_arrival_workbook(content, default_store_id)
    reconcile_batches(repository, batches)
    return repository.save_arrival_import(
        filename=filename,
        file_hash=file_hash,
        default_store_id=default_store_id,
        batches=batches,
    )


def reconcile_saved_arrivals(repository: Any) -> dict[str, int]:
    batches = repository.get_arrival_batches()
    if not batches:
        return {"matched": 0, "unmatched": 0, "conflict": 0}
    reconcile_batches(repository, batches)
    repository.update_arrival_matches(batches)
    counts = defaultdict(int)
    for batch in batches:
        for item in batch["items"]:
            counts[item["match_status"]] += 1
    return {
        "matched": counts["matched"],
        "unmatched": counts["unmatched"],
        "conflict": counts["conflict"],
    }


def expected_receive_from_tracking(batch: dict[str, Any]) -> str | None:
    if batch.get("actual_receive_date"):
        return _text(batch["actual_receive_date"])
    if batch.get("expected_receive_date"):
        return _text(batch["expected_receive_date"])
    # The safety setting already covers receiving uncertainty, so tracking
    # dates are used directly without adding another FBT receiving buffer.
    signed_date = (
        batch.get("port_arrival_date")
        or batch.get("actual_signed_date")
        or batch.get("expected_signed_date")
    )
    if not signed_date:
        return None
    return date.fromisoformat(_text(signed_date)).isoformat()
