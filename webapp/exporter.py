from __future__ import annotations

import math
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_SHIPPING_HEADERS = [
    "品名",
    "MSKU",
    "7天销量",
    "14天销量",
    "30天销量",
    "预测日销量",
    "FBT库存",
    "FBT可售",
    "FBT在途",
    "安全天数",
    "本地仓发货频率",
]


def _active_shipping_channels(
    dashboard: dict[str, Any],
) -> list[dict[str, Any]]:
    products = dashboard.get("products", [])
    if not products:
        return []
    channels = [
        plan
        for plan in products[0].get("channel_plans", [])
        if plan.get("enabled", True)
    ]
    return sorted(
        channels,
        key=lambda channel: str(
            channel.get("planning_arrival_date")
            or channel.get("arrival_date")
            or ""
        ),
    )


def _channel_coverage_days(
    item: dict[str, Any],
    plan: dict[str, Any],
    channels: list[dict[str, Any]],
) -> float:
    regular_channels = [
        channel
        for channel in channels
        if channel.get("key") not in {"express", "air"}
    ]
    seasonal_channel = max(
        regular_channels,
        key=lambda channel: str(
            channel.get("planning_arrival_date")
            or channel.get("arrival_date")
            or ""
        ),
        default=None,
    )
    if seasonal_channel and seasonal_channel.get("key") == plan.get("key"):
        return float(item.get("current_total_coverage_days") or 0)
    return float(
        plan.get("target_coverage_days")
        or plan.get("planning_arrival_days")
        or 0
    )


def build_export(
    dashboard: dict[str, Any], shipment_data: dict[str, Any] | None = None
) -> bytes:
    channels = _active_shipping_channels(dashboard)
    headers = [
        *BASE_SHIPPING_HEADERS,
        *(str(channel["label"]) for channel in channels),
        *(f"{channel['label']}建议补货" for channel in channels),
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "补货建议"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    export_products = [
        item
        for item in dashboard["products"]
        if not item.get("is_planning_excluded")
    ]
    for row_index, item in enumerate(export_products, start=2):
        values = [
            item["product_name"],
            item["msku"],
            round(float(item["avg_7"]), 2),
            round(float(item["avg_14"]), 2),
            round(float(item["avg_30"]), 2),
            round(float(item["dynamic_daily"]), 2),
            item["fbt_total"],
            item["fbt_sellable"],
            item["fbt_in_transit"],
            item["safety_buffer_days"],
            item["dispatch_interval_days"],
            *(
                round(_channel_coverage_days(item, channel, channels), 4)
                for channel in channels
            ),
            *(
                item.get(f"{channel['key']}_qty", 0)
                for channel in channels
            ),
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column, value=value)

    for index, header in enumerate(headers, start=1):
        width = 34 if header == "品名" else 24 if header == "MSKU" else 18
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_purchase_export(purchase_plan: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(["SKU", "品名", "备货量"])

    thin = Side(style="thin", color="60656F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for item in purchase_plan.get("items", []):
        if float(item.get("final_qty") or 0) <= 0:
            continue
        sheet.append(
            [
                item.get("sku", ""),
                item.get("product_name", ""),
                int(math.ceil(float(item.get("final_qty") or 0))),
            ]
        )

    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 52
    sheet.column_dimensions["C"].width = 14
    for row in sheet.iter_rows(
        min_row=1, max_row=max(1, sheet.max_row), min_col=1, max_col=3
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if cell.column == 3 else "left",
                vertical="center",
            )
        row[2].number_format = "#,##0"
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.auto_filter.ref = f"A1:C{max(1, sheet.max_row)}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


TRACKING_SYSTEM_HEADERS = [
    "record_id",
    "batch_id",
    "cargo_code",
    "store_id",
    "batch_label",
    "shipment_date",
    "departure_date",
    "port_arrival_date",
    "expected_signed_date",
    "actual_signed_date",
    "expected_receive_date",
    "actual_receive_date",
    "is_fully_received",
    "carrier",
    "tracking_number",
    "status_note",
    "route_note",
    "metric_f",
    "metric_g",
    "metric_h",
    "item_id",
    "raw_sku",
    "matched_store_id",
    "matched_msku",
    "matched_sku",
    "product_name",
    "shipment_qty",
    "match_status",
    "match_method",
    "conflict_note",
    "source_row",
]


def _excel_date(value: Any) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _status_note(shipment: dict[str, Any]) -> str:
    parts: list[str] = []
    source_note = str(shipment.get("status_note") or "").rstrip("；;")
    if source_note:
        parts.append(source_note)
    if not source_note and shipment.get("expected_signed_date"):
        parts.append(f"预计{shipment['expected_signed_date'][5:]}签收")
    if not source_note and shipment.get("actual_signed_date"):
        parts.append(f"{shipment['actual_signed_date'][5:]}已签收")
    if shipment.get("is_received") and "已全部接收" not in source_note:
        receive_date = shipment.get("actual_receive_date")
        parts.append(
            f"{receive_date[5:]}已全部接收" if receive_date else "已全部接收"
        )
    elif shipment.get("received_qty", 0) > 0 and "已接收" not in source_note:
        parts.append(f"FBT已接收{shipment['received_qty']:g}件")
    elif (
        shipment.get("awaiting_receive_qty", 0) > 0
        and "待FBT接收" not in source_note
    ):
        parts.append(f"{shipment['awaiting_receive_qty']:g}件待FBT接收")
    return "；".join(dict.fromkeys(part for part in parts if part)) + (
        "；" if parts else ""
    )


def _route_note(shipment: dict[str, Any]) -> str:
    if shipment.get("route_note"):
        return str(shipment["route_note"])
    parts = []
    if shipment.get("departure_date"):
        parts.append(f"{shipment['departure_date'][5:]}开船")
    if shipment.get("port_arrival_date"):
        parts.append(f"{shipment['port_arrival_date'][5:]}到港")
    return " ".join(parts)


def _style_tracking_sheet(sheet: Any, max_row: int) -> None:
    thin = Side(style="thin", color="60656F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.sheet_view.showGridLines = False
    widths = [13, 34, 20, 34, 12, 11, 11, 22, 42, 28]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=10):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                vertical="center",
                horizontal="left",
                wrap_text=True,
            )
            cell.font = Font(name="Microsoft YaHei", size=10)
        row[4].alignment = Alignment(horizontal="right", vertical="center")
    sheet.freeze_panes = "C1"


def build_arrival_tracking_export(shipment_data: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "到货跟踪"

    system_rows: list[list[Any]] = []
    detail_rows: list[list[Any]] = []
    issue_ids = {
        item["item_id"] for item in shipment_data.get("reconciliation_issues", [])
    }
    current_row = 1
    date_groups: dict[str, list[int]] = {}
    source_shipments = shipment_data.get("shipments", [])
    imported_dates = [
        str(item.get("delivery_time") or item.get("create_time"))[:10]
        for item in source_shipments
        if item.get("arrival_batch_id")
        and (item.get("delivery_time") or item.get("create_time"))
    ]
    tracking_start_date = min(imported_dates) if imported_dates else None
    export_shipments = [
        item
        for item in source_shipments
        if float(item.get("shipment_qty") or 0) > 0
        and (
            item.get("arrival_batch_id")
            or not tracking_start_date
            or str(item.get("delivery_time") or item.get("create_time") or "")[:10]
            >= tracking_start_date
        )
    ]
    ordered = sorted(
        export_shipments,
        key=lambda item: (
            item.get("delivery_time") or item.get("create_time") or "9999-12-31",
            item.get("cargo_code") or "",
        ),
    )
    for shipment in ordered:
        items = [
            item
            for item in shipment.get("items", [])
            if float(item.get("shipment_qty") or 0) > 0
        ] or [
            {
                "msku": "",
                "sku": "",
                "raw_sku": "",
                "product_name": "",
                "shipment_qty": shipment.get("shipment_qty", 0),
            }
        ]
        start_row = current_row
        shipment_date = shipment.get("delivery_time") or shipment.get("create_time")
        batch_id = shipment.get("arrival_batch_id") or (
            f"ibr_{shipment['cargo_code']}"
        )
        batch_label = shipment.get("batch_label") or shipment.get("cargo_code", "")
        status_note = _status_note(shipment)
        route_note = _route_note(shipment)
        for item in items:
            raw_sku = item.get("raw_sku") or item.get("msku") or item.get("sku")
            item_id = item.get("arrival_item_id") or (
                f"api_{shipment.get('cargo_code', '')}_{item.get('msku', raw_sku)}"
            )
            row = [
                _excel_date(shipment_date),
                batch_label,
                raw_sku,
                item.get("product_name", ""),
                item.get("shipment_qty", 0),
                shipment.get("metric_f"),
                shipment.get("metric_g"),
                shipment.get("metric_h"),
                status_note,
                route_note,
            ]
            sheet.append(row)
            detail_rows.append(
                [
                    shipment.get("status_name", ""),
                    shipment.get("cargo_code", ""),
                    shipment.get("store_name", ""),
                    raw_sku,
                    item.get("msku", ""),
                    item.get("sku", ""),
                    item.get("product_name", ""),
                    item.get("shipment_qty", 0),
                    item.get("signed_qty", 0),
                    item.get("received_qty", 0),
                    item.get("remaining_qty", 0),
                    shipment.get("expected_signed_date"),
                    shipment.get("expected_receive_date"),
                    shipment.get("carrier", ""),
                    shipment.get("tracking_number", ""),
                    item.get("match_status") or "matched",
                    item.get("match_method") or "领星API",
                    "是" if item_id in issue_ids else "",
                ]
            )
            system_rows.append(
                [
                    f"{batch_id}:{item_id}",
                    batch_id,
                    shipment.get("cargo_code", ""),
                    shipment.get("store_id", ""),
                    batch_label,
                    shipment_date,
                    shipment.get("departure_date"),
                    shipment.get("port_arrival_date"),
                    shipment.get("expected_signed_date"),
                    shipment.get("actual_signed_date"),
                    shipment.get("expected_receive_date"),
                    shipment.get("actual_receive_date"),
                    bool(shipment.get("is_received")),
                    shipment.get("carrier", ""),
                    shipment.get("tracking_number", ""),
                    status_note,
                    route_note,
                    shipment.get("metric_f"),
                    shipment.get("metric_g"),
                    shipment.get("metric_h"),
                    item_id,
                    raw_sku,
                    item.get("store_id") or shipment.get("store_id", ""),
                    item.get("msku", ""),
                    item.get("sku", ""),
                    item.get("product_name", ""),
                    item.get("shipment_qty", 0),
                    item.get("match_status") or "matched",
                    item.get("match_method") or "领星API",
                    "",
                    current_row,
                ]
            )
            current_row += 1
        end_row = current_row - 1
        for column in (2, 6, 7, 8, 9, 10):
            if end_row > start_row:
                sheet.merge_cells(
                    start_row=start_row,
                    start_column=column,
                    end_row=end_row,
                    end_column=column,
                )
        if shipment_date:
            date_groups.setdefault(str(shipment_date)[:10], []).extend(
                [start_row, end_row]
            )

    for rows in date_groups.values():
        start_row, end_row = min(rows), max(rows)
        if end_row > start_row:
            sheet.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=end_row,
                end_column=1,
            )
    _style_tracking_sheet(sheet, max(1, sheet.max_row))
    for cell in sheet["A"]:
        cell.number_format = "m/d/yy"
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in sheet["H"]:
        if cell.value not in (None, ""):
            cell.fill = PatternFill("solid", fgColor="39CF28")

    detail = workbook.create_sheet("对账明细")
    detail_headers = [
        "状态",
        "IBR",
        "店铺",
        "原始SKU",
        "匹配MSKU",
        "平台SKU",
        "产品名",
        "发货量",
        "签收量",
        "FBT接收量",
        "未完成入库量",
        "预计签收日",
        "预计FBT入库日",
        "承运商",
        "跟踪号",
        "匹配状态",
        "匹配方式",
        "待处理",
    ]
    detail.append(detail_headers)
    for row in detail_rows:
        detail.append(row)
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:R{max(1, detail.max_row)}"
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")
    for index, header in enumerate(detail_headers, start=1):
        detail.column_dimensions[get_column_letter(index)].width = (
            30
            if header in {"产品名", "跟踪号"}
            else 23
            if header in {"IBR", "原始SKU", "匹配MSKU"}
            else 15
        )

    issues = workbook.create_sheet("异常待处理")
    issue_headers = [
        "IBR",
        "店铺ID",
        "原始SKU",
        "产品名",
        "发货量",
        "问题",
        "系统明细ID",
    ]
    issues.append(issue_headers)
    for item in shipment_data.get("reconciliation_issues", []):
        issues.append(
            [
                item["cargo_code"],
                item["store_id"],
                item["raw_sku"],
                item["product_name"],
                item["shipment_qty"],
                item["conflict_note"],
                item["item_id"],
            ]
        )
    issues.freeze_panes = "A2"
    issues.auto_filter.ref = f"A1:G{max(1, issues.max_row)}"
    for cell in issues[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="DC2626")
    for index, width in enumerate([24, 20, 22, 34, 12, 44, 28], start=1):
        issues.column_dimensions[get_column_letter(index)].width = width

    system = workbook.create_sheet("_系统数据")
    system.append(TRACKING_SYSTEM_HEADERS)
    for row in system_rows:
        system.append(row)
    system.sheet_state = "hidden"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
